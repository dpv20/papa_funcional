# funciones/Trans_excel.py
from pathlib import Path
import os
import platform
import subprocess
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

# ---------------- estilos ----------------
THIN = Side(style="thin")
ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)
BOLD_U = Font(bold=True, underline="single")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NOWRAP = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# Formatos numéricos (comma style)
FMT_QTY = '#,##0'       # Cantidades sin decimales
FMT_MONEY = '#,##0'     # Precios/Totales sin decimales
FMT_GENERIC = '#,##0.00'

# Orden sugerido de tipos
TYPE_ORDER = [
    "MATERIALES",
    "EQUIPOS",
    "MAQUINARIAS",
    "HERRAMIENTAS",
    "MANO DE OBRA",
    "SERVICIOS",
]

# ---------------- columnas ----------------
def set_column_widths(ws):
    """
    Tamaños base:
      A: 0.25x usual
      B: 1x usual
      C: 4x usual
      D: 0.5x usual
      E,F,G: 1x usual
    """
    usual = 8.43
    widths = {"A": usual * 0.25, "B": usual, "C": usual * 4, "D": usual * 0.5,
              "E": usual, "F": usual, "G": usual}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def scale_columns(ws, scales: dict[str, float]):
    """Multiplica el ancho actual de las columnas indicadas por un factor."""
    for col, factor in scales.items():
        cur = ws.column_dimensions[col].width
        if cur is None:
            continue
        ws.column_dimensions[col].width = cur * float(factor)

def draw_all_borders(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = ALL_THIN

def set_num(cell, val, fmt=FMT_MONEY, align=RIGHT, bold=False):
    cell.value = float(val or 0)
    cell.number_format = fmt
    cell.alignment = align
    if bold:
        cell.font = BOLD

# ---------------- mapeo de tipos ----------------
def _load_tipo_mapping(base: Path) -> dict:
    """
    Lee 'categorias.csv' si existe.
    Soporta:
      - ['Codigo','Tipo'] (mapeo por código)
      - ['Categoria','Tipo'] (mapeo por categoría del maestro)
    Retorna {'by_code': {...}, 'by_category': {...}} o {}.
    """
    path = base / "categorias.csv"
    if not path.exists():
        return {}

    df = pd.read_csv(path)
    mapping = {}

    def real_name(name):
        for c in df.columns:
            if c.strip().lower() == name:
                return c
        return None

    tipo_col = real_name("tipo")
    if tipo_col is None:
        return {}

    codigo_col = real_name("codigo")
    categoria_col = real_name("categoria")

    if codigo_col is not None:
        tmp = df[[codigo_col, tipo_col]].dropna()
        mapping["by_code"] = dict(zip(tmp[codigo_col].astype(str), tmp[tipo_col].astype(str)))

    if categoria_col is not None:
        tmp = df[[categoria_col, tipo_col]].dropna()
        mapping["by_category"] = dict(zip(tmp[categoria_col].astype(str), tmp[tipo_col].astype(str)))

    return mapping

def _tipo_from_row(codigo: str, categoria: str, mapping: dict) -> str:
    """
    Determina el 'Tipo' de una fila usando el mapping (prioriza por Código).
    Si no encuentra, usa la categoría si coincide con TYPE_ORDER; si no, 'MATERIALES'.
    """
    if mapping:
        if "by_code" in mapping:
            t = mapping["by_code"].get(str(codigo))
            if t:
                return str(t).strip().upper()
        if "by_category" in mapping:
            t = mapping["by_category"].get(str(categoria))
            if t:
                return str(t).strip().upper()

    cat_up = (str(categoria) if pd.notna(categoria) else "").strip().upper()
    if cat_up in TYPE_ORDER:
        return cat_up
    return "MATERIALES"

# ---------------- escritura ----------------
def _write_headers(ws, start_row: int, item_row: pd.Series) -> int:
    """
    Escribe cabecera de un ítem y retorna el siguiente row disponible
    (dejando UNA fila en blanco después del encabezado de tabla).
    También aplica el escalado adicional de columnas.
    """
    row = start_row
    # Título principal
    ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=7)
    c = ws.cell(row=row+1, column=2, value="ANÁLISIS DE PRECIOS UNITARIOS")
    c.font = BOLD_U
    c.alignment = CENTER

    row += 3
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=f"ITEM: {item_row['Item']}").font = BOLD
    ws.cell(row=row, column=6, value=f"CANTIDAD ({item_row['cantidad tipo']}):").font = BOLD
    set_num(ws.cell(row=row, column=7), item_row['cantidad numero'], fmt=FMT_QTY, bold=True)

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value=f"PARTIDA: {item_row['Partida']}").font = BOLD
    ws.cell(row=row, column=6, value="MONEDA:").font = BOLD
    # MONEDA como texto (por si es 'CLP', 'USD', etc.)
    moneda_val = str(item_row.get('moneda', '') or '')
    cmon = ws.cell(row=row, column=7, value=moneda_val)
    cmon.alignment = RIGHT
    cmon.font = BOLD

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=f"FECHA: {item_row['Fecha']}").font = BOLD

    # Encabezados de la tabla
    row += 2
    headers = ["ITEM", "DESCRIPCIÓN", "UD", "CANTIDAD", "P. UNITARIO", "TOTAL"]
    for j, text in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=j, value=text)
        cell.font = BOLD
        cell.alignment = CENTER_NOWRAP  # no agrandar altura
    draw_all_borders(ws, f"B{row}:G{row}")

    # 1) Fila en blanco solicitada tras el encabezado
    row += 1
    # 2) Escalado adicional solicitado
    #    << AQUÍ se modifican tamaños después del encabezado >>
    scale_columns(ws, {"B": 1.1, "E": 1.1, "F": 1.1, "G": 1.1})

    # Siguiente fila disponible
    return row + 1

def _write_tipo_block(ws, start_row: int, det_tipo: pd.DataFrame, tipo_name: str) -> tuple[int, float]:
    """
    Escribe bloque para un 'tipo' específico (si hay filas).
    Retorna (siguiente_row, subtotal_tipo).
    """
    if det_tipo.empty:
        return start_row, 0.0

    # Título del bloque (fila actual)
    row = start_row
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    title = ws.cell(row=row, column=2, value=tipo_name)
    title.font = BOLD
    draw_all_borders(ws, f"B{row}:G{row}")

    subtotal = 0.0
    # Filas de detalle
    row += 1
    det_tipo = det_tipo.copy()
    det_tipo["Codigo"] = det_tipo["Codigo"].astype(str)
    det_tipo = det_tipo.sort_values("Codigo")

    for _, r in det_tipo.iterrows():
        codigo = r["Codigo"]
        desc = r.get("Resumen", "")
        ud = r.get("Ud", "")
        cant = float(r.get("cantidad", 0) or 0)
        unit = float(r.get("Pres", 0) or 0)
        total = cant * unit
        subtotal += total

        ws.cell(row=row, column=2, value=codigo)
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=4, value=ud)

        set_num(ws.cell(row=row, column=5), cant, fmt=FMT_QTY)       # Cantidad (sin dec)
        set_num(ws.cell(row=row, column=6), unit, fmt=FMT_MONEY)     # P. Unitario
        set_num(ws.cell(row=row, column=7), total, fmt=FMT_MONEY)    # Total

        draw_all_borders(ws, f"B{row}:G{row}")
        row += 1

    # Subtotal del bloque (sin multiplicar por cantidad número)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    stc = ws.cell(row=row, column=2, value=f"Subtotal {tipo_name.title()}")
    stc.font = BOLD
    set_num(ws.cell(row=row, column=7), subtotal, fmt=FMT_MONEY, bold=True)
    draw_all_borders(ws, f"B{row}:G{row}")

    # Deja UNA fila en blanco después del bloque
    return row + 2, subtotal

# ---------------- generación de Excel ----------------
def generar_excel(proyecto: str,
                  base_dir: str = ".",
                  maestro_csv: str = "construction_budget_data.csv",
                  salida: str | Path | None = None) -> Path:
    """
    Genera un Excel APU para un proyecto, agrupando por Tipo (según categorias.csv).
    Para cada ítem:
      - Bloques por tipo presentes (MATERIALES, EQUIPOS, MAQUINARIAS, HERRAMIENTAS, MANO DE OBRA, SERVICIOS).
      - Subtotal por bloque (sin multiplicar por 'cantidad numero').
      - 'Precio Unitario' (suma de subtotales) con 1 fila en blanco antes.
      - 'TOTAL PARTIDA (Ud)' = Precio Unitario * cantidad numero, con 1 fila en blanco antes.
    Retorna la ruta del archivo generado.
    """
    base = Path(base_dir)
    proyecto_dir = base / "presupuestos" / proyecto
    maestro_csv = base / maestro_csv

    datos_csv = proyecto_dir / "datos.csv"
    detalle_csv = proyecto_dir / "detalle.csv"
    if not datos_csv.exists() or not detalle_csv.exists():
        raise FileNotFoundError("No se encuentran datos.csv o detalle.csv en el proyecto seleccionado")

    df_datos = pd.read_csv(datos_csv)
    df_detalle = pd.read_csv(detalle_csv)
    df_maestro = pd.read_csv(maestro_csv)

    # Cargar mapeo de tipos desde categorias.csv (si existe)
    tipo_mapping = _load_tipo_mapping(base)

    # --- Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "APU"
    # << AQUÍ se fijan los tamaños base de columnas >>
    set_column_widths(ws)

    row = 1
    for _, item_row in df_datos.iterrows():
        # Cabecera
        row = _write_headers(ws, row, item_row)

        # Detalle del ítem actual
        det_item = df_detalle[df_detalle["item"].astype(str) == str(item_row["Item"])].copy()

        if det_item.empty:
            # Precio Unitario
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            ws.cell(row=row, column=5, value="Precio Unitario").font = BOLD
            set_num(ws.cell(row=row, column=7), 0, fmt=FMT_MONEY, bold=True)
            draw_all_borders(ws, f"E{row}:G{row}")

            # Fila en blanco
            row += 1

            # TOTAL PARTIDA (Ud)
            row += 1
            unidad = str(item_row.get("cantidad tipo", "") or "").strip()
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            ws.cell(row=row, column=5, value=f"TOTAL PARTIDA ({unidad})").font = BOLD
            set_num(ws.cell(row=row, column=7), 0, fmt=FMT_MONEY, bold=True)
            draw_all_borders(ws, f"E{row}:G{row}")

            # Separación antes del siguiente ítem
            row += 3
            continue

        # Merge con maestro para obtener Resumen, Ud, Pres, Categoria
        det_item["Codigo"] = det_item["Codigo"].astype(str)
        df_maestro["Codigo"] = df_maestro["Codigo"].astype(str)
        det_full = det_item.merge(
            df_maestro[["Codigo", "Resumen", "Ud", "Pres", "Categoria"]],
            on="Codigo", how="left"
        )

        # Asignar Tipo usando categorias.csv (o inferir)
        det_full["Tipo"] = det_full.apply(
            lambda r: _tipo_from_row(r["Codigo"], r.get("Categoria", ""), tipo_mapping),
            axis=1
        )

        # Escribir bloques por cada Tipo presente y acumular subtotales
        subtotales = []
        presentes = set(det_full["Tipo"].astype(str).str.upper())
        tipos_presentes = [t for t in TYPE_ORDER if t in presentes]
        otros_tipos = sorted(presentes - set(TYPE_ORDER))
        for tipo in tipos_presentes + otros_tipos:
            det_tipo = det_full[det_full["Tipo"].astype(str).str.upper() == tipo]
            row, sub_t = _write_tipo_block(ws, row, det_tipo, tipo)
            if sub_t > 0:
                subtotales.append(sub_t)

        # Precio Unitario = suma de subtotales
        precio_unitario = float(sum(subtotales)) if subtotales else 0.0

        # Precio Unitario (ya venimos con una fila en blanco del último bloque)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value="Precio Unitario").font = BOLD
        set_num(ws.cell(row=row, column=7), precio_unitario, fmt=FMT_MONEY, bold=True)
        draw_all_borders(ws, f"E{row}:G{row}")

        # Fila en blanco antes de TOTAL PARTIDA
        row += 1

        # TOTAL PARTIDA (Ud) = Precio Unitario * cantidad numero
        row += 1
        unidad = str(item_row.get("cantidad tipo", "") or "").strip()
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=f"TOTAL PARTIDA ({unidad})").font = BOLD
        total_partida = precio_unitario * float(item_row.get("cantidad numero", 0) or 0)
        set_num(ws.cell(row=row, column=7), total_partida, fmt=FMT_MONEY, bold=True)
        draw_all_borders(ws, f"E{row}:G{row}")

        # Separación antes del siguiente ítem
        row += 2

    if salida is None:
        salida = proyecto_dir / "presupuesto_APU.xlsx"
    else:
        salida = Path(salida)

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return Path(salida)

# ---------------- helpers reutilización/visualización ----------------
def excel_output_path(proyecto: str, base_dir: str = ".") -> Path:
    """Ruta donde se guarda el Excel del proyecto."""
    return Path(base_dir) / "presupuestos" / proyecto / "presupuesto_APU.xlsx"

def obtener_o_generar_excel(
    proyecto: str,
    base_dir: str = ".",
    maestro_csv: str = "construction_budget_data.csv",
    reescribir: bool = False
) -> Path:
    """
    Si el Excel existe y 'reescribir' es False: devuelve la ruta sin modificarlo.
    Si no existe, lo genera.
    Si 'reescribir' es True, lo regenera y reemplaza el anterior.
    """
    out_path = excel_output_path(proyecto, base_dir)
    if out_path.exists() and not reescribir:
        return out_path
    return generar_excel(proyecto=proyecto, base_dir=base_dir, maestro_csv=maestro_csv, salida=out_path)

def _abrir_en_sistema(path: Path) -> bool:
    """Abre el archivo con la aplicación por defecto del sistema operativo."""
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(str(path))  # type: ignore[attr-defined]
            return True
        elif system == "Darwin":  # macOS
            subprocess.Popen(["open", str(path)])
            return True
        else:  # Linux / otros
            subprocess.Popen(["xdg-open", str(path)])
            return True
    except Exception:
        return False

def ver_excel(
    proyecto: str,
    base_dir: str = ".",
    maestro_csv: str = "construction_budget_data.csv",
    reescribir: bool = False
) -> Path:
    """
    Abre el Excel existente (o lo genera si no existe o si reescribir=True) con la
    app por defecto del sistema. Retorna la ruta del archivo.
    """
    path = obtener_o_generar_excel(
        proyecto=proyecto,
        base_dir=base_dir,
        maestro_csv=maestro_csv,
        reescribir=reescribir
    )
    _abrir_en_sistema(path)
    return path

# ---------------- Vista Streamlit: Crear/Mostrar/Descargar ----------------
def _listar_proyectos(base_dir: str = ".") -> list[str]:
    base = Path(base_dir) / "presupuestos"
    if not base.exists():
        return []
    proyectos = []
    for p in base.iterdir():
        if p.is_dir() and (p / "datos.csv").exists() and (p / "detalle.csv").exists():
            proyectos.append(p.name)
    return sorted(proyectos)

def render():
    """Vista 'Crear Excel' con tres botones: Generar, Mostrar, Descargar."""
    st.title("📄 Crear/Ver/Descargar Excel APU")

    proyectos = _listar_proyectos(".")
    if not proyectos:
        st.info("No hay proyectos con `datos.csv` y `detalle.csv` en `presupuestos/`.")
        return

    proyecto = st.selectbox("Proyecto", proyectos, index=0)
    ruta_xlsx = excel_output_path(proyecto)
    st.caption(f"Archivo de salida: `{ruta_xlsx}`")

    col1, col2, col3 = st.columns(3)

    # Generar (forzar)
    with col1:
        if st.button("⚙️ Generar Excel", use_container_width=True):
            try:
                path = obtener_o_generar_excel(proyecto, reescribir=True)
                st.success(f"Excel generado: {path.name}")
            except Exception as e:
                st.error(f"Error al generar: {e}")

    # Mostrar (abrir; si no existe, generar)
    with col2:
        if st.button("👁️ Mostrar Excel", use_container_width=True):
            try:
                ver_excel(proyecto, reescribir=False)  # genera si falta
                st.success("Abriendo Excel...")
            except Exception as e:
                st.error(f"No se pudo abrir: {e}")

    # Descargar (si no existe, generar)
    with col3:
        # Para no requerir doble click, preparamos el buffer al renderizar el botón
        try:
            path_ready = obtener_o_generar_excel(proyecto, reescribir=False)
            with open(path_ready, "rb") as f:
                data = f.read()
        except Exception as e:
            data = None
            st.error(f"No se pudo preparar el archivo para descarga: {e}")

        st.download_button(
            "⬇️ Descargar Excel",
            data=data if data is not None else b"",
            file_name=f"{proyecto}_presupuesto_APU.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=(data is None),
            use_container_width=True,
        )

    # Estado actual
    if ruta_xlsx.exists():
        st.caption("Estado: ✅ Existe")
    else:
        st.caption("Estado: ❌ No existe (se generará automáticamente al usar **Mostrar** o **Descargar**).")

# Alias por si tu app.py invoca otro nombre
def render_crear_excel():
    render()
