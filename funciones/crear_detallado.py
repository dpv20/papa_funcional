# funciones/crear_detallado.py
"""
Genera Excel de Presupuesto Detallado con formato:
- Header: PRESUPUESTO DETALLADO, PROYECTO, UBICACIÓN, NOMBRE PROPIETARIO, FECHA
- Tabla: ITEM, DESCRIPCIÓN, UD, CANTIDAD, P. UNITARIO, TOTAL
- Agrupación por ítems padre con subtotales
- Footer: TOTAL COSTO DIRECTO
"""
from pathlib import Path
import os
import platform
import subprocess
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .monedas import get_moneda_value

# ---------------- Estilos ----------------
THIN = Side(style="thin")
ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)
BOLD_U = Font(bold=True, underline="single")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

# Formato numérico con locale español (punto para miles, coma para decimales)
FMT_QTY = '[$-340A]#,##0.00'
FMT_MONEY = '[$-340A]#,##0.00'


def _parse_key(code: str) -> Tuple[int, ...]:
    """Ordena jerárquicamente: 1 < 1.01 < 1.02 < 2 < 2.01"""
    parts = []
    for seg in str(code).split("."):
        try:
            parts.append(int(seg))
        except Exception:
            parts.append(0)
    return tuple(parts)


def _collect_parents(items: List[str]) -> List[str]:
    """Devuelve padres que tienen hijos: '1', '2', etc."""
    items = [str(x) for x in items if isinstance(x, str)]
    children_by_parent: Dict[str, List[str]] = {}
    for it in items:
        segs = it.split(".")
        if len(segs) >= 2:
            parent = segs[0]  # Solo nivel superior
            children_by_parent.setdefault(parent, []).append(it)
    parents = [p for p, childs in children_by_parent.items() if len(childs) > 0]
    return sorted(set(parents), key=_parse_key)


def _children_of_parent(items: List[str], parent: str) -> List[str]:
    """Hijos directos de un padre (ej: 1 -> [1.01, 1.02])"""
    pref = parent + "."
    childs = [it for it in items if it.startswith(pref)]
    return sorted(childs, key=_parse_key)


def _listar_proyectos(base_dir: str = ".") -> list[str]:
    base = Path(base_dir) / "presupuestos"
    if not base.exists():
        return []
    proyectos = []
    for p in base.iterdir():
        if p.is_dir() and (p / "datos.csv").exists() and (p / "detalle.csv").exists():
            proyectos.append(p.name)
    return sorted(proyectos)


def _abrir_en_sistema(path: Path) -> bool:
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(str(path))
            return True
        elif system == "Darwin":
            subprocess.Popen(["open", str(path)])
            return True
        else:
            subprocess.Popen(["xdg-open", str(path)])
            return True
    except Exception:
        return False


def _compute_precio_unitario_por_item(proyecto_dir: Path, df_maestro: pd.DataFrame) -> Dict[str, float]:
    """Precio Unitario por Item = sum(cantidad * Pres) de su detalle."""
    detalle_csv = proyecto_dir / "detalle.csv"
    if not detalle_csv.exists():
        return {}
    df_detalle = pd.read_csv(detalle_csv)
    if df_detalle.empty:
        return {}

    df_detalle["item"] = df_detalle["item"].astype(str)
    df_detalle["Codigo"] = df_detalle["Codigo"].astype(str)
    df_maestro = df_maestro.copy()
    df_maestro["Codigo"] = df_maestro["Codigo"].astype(str)

    det_full = df_detalle.merge(
        df_maestro[["Codigo", "Pres"]],
        on="Codigo", how="left"
    )
    det_full["cantidad"] = pd.to_numeric(det_full["cantidad"], errors="coerce").fillna(0.0)
    det_full["Pres"] = pd.to_numeric(det_full["Pres"], errors="coerce").fillna(0.0)
    det_full["subtotal_linea"] = det_full["cantidad"] * det_full["Pres"]

    precios = det_full.groupby("item")["subtotal_linea"].sum().to_dict()
    return {str(k): float(v) for k, v in precios.items()}


    # ... (Signature update) ...
def generar_excel_detallado(
    proyecto: str,
    base_dir: str = ".",
    maestro_csv: str = "construction_budget_data.csv",
    nombres_padres: Dict[str, str] | None = None,
    ubicacion: str = "",
    propietario: str = "",
    porc_utilidad: float = 0.0,
    porc_iva: float = 19.0,
    salida: str | Path | None = None
) -> Path:
    # ... (existing setup code) ...
    base = Path(base_dir)
    proyecto_dir = base / "presupuestos" / proyecto
    maestro_csv_path = base / maestro_csv
    datos_csv = proyecto_dir / "datos.csv"
    detalle_csv = proyecto_dir / "detalle.csv"

    if not datos_csv.exists() or not detalle_csv.exists():
        raise FileNotFoundError("No se encuentran datos.csv o detalle.csv del proyecto.")

    # IMPORTANTE: dtype={"Item": str} para evitar que 1.01 sea float
    df_datos = pd.read_csv(datos_csv, dtype={"Item": str})
    df_maestro = pd.read_csv(maestro_csv_path)

    moneda_proyecto = "CLP"
    if not df_datos.empty:
        row_101 = df_datos[df_datos["Item"] == "1.01"]
        if not row_101.empty and "moneda" in row_101.columns:
            moneda_proyecto = str(row_101.iloc[0]["moneda"] or "CLP")
        elif "moneda" in df_datos.columns:
            moneda_proyecto = str(df_datos["moneda"].iloc[0] or "CLP")
    
    factor_conversion = get_moneda_value(moneda_proyecto)

    items = sorted(df_datos["Item"].astype(str).tolist(), key=_parse_key)
    parents = _collect_parents(items)
    nombres_padres = nombres_padres or {}

    fecha_str = ""
    if not df_datos.empty and "Fecha" in df_datos.columns:
         fecha_str = str(df_datos["Fecha"].iloc[0] or "")

    precios_item = _compute_precio_unitario_por_item(proyecto_dir, df_maestro)
    df_datos_idx = df_datos.set_index("Item")

    # --- Crear Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detallado ({moneda_proyecto})"

    # --- Configurar Grid (Columnas B a S) ---
    ws.column_dimensions["A"].width = 2
    for c_idx in range(2, 20): 
        col_letter = get_column_letter(c_idx)
        ws.column_dimensions[col_letter].width = 5

    # Funciones Helper para Merged Cells
    def write_merged(r, c_start, span, val, style=None, align=None, number_format=None, border=None, target_ws=None):
        working_ws = target_ws if target_ws else ws
        c_end = c_start + span - 1
        cell = working_ws.cell(row=r, column=c_start, value=val)
        working_ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
        if style:
            cell.font = style
        if align:
            cell.alignment = align
        if number_format:
            cell.number_format = number_format
        
        if border:
            for i in range(c_start, c_end + 1):
                 working_ws.cell(row=r, column=i).border = border
        
        return cell

    # Mapeo de Columnas (Start Col, Span)
    COL_ITEM = (2, 1)   # B
    COL_DESC = (3, 7)   # C-I
    COL_UD   = (10, 1)  # J
    COL_CANT = (11, 2)  # K-L
    COL_PU   = (13, 3)  # M-O
    COL_TOT  = (16, 4)  # P-S

    row = 2

    # --- Titulo Principal ---
    write_merged(row, 2, 18, "PRESUPUESTO DETALLADO", style=BOLD_U, align=CENTER)
    row += 2

    # --- Header Proyecto ---
    write_merged(row, 2, 3, "PROYECTO:", style=BOLD)
    write_merged(row, 5, 8, proyecto)
    
    write_merged(row, 13, 3, "FECHA:", style=BOLD, align=RIGHT)
    write_merged(row, 16, 4, fecha_str, align=CENTER)
    row += 1

    write_merged(row, 2, 3, "UBICACIÓN:", style=BOLD)
    write_merged(row, 5, 8, ubicacion)
    
    write_merged(row, 13, 3, "MONEDA:", style=BOLD, align=RIGHT)
    write_merged(row, 16, 4, moneda_proyecto, align=CENTER)
    row += 1
    
    write_merged(row, 2, 3, "PROPIETARIO:", style=BOLD)
    write_merged(row, 5, 8, propietario)
    row += 2

    # --- Encabezados Tabla (CON BORDES) ---
    write_merged(row, *COL_ITEM, "ITEM", style=BOLD, align=CENTER, border=ALL_THIN)
    write_merged(row, *COL_DESC, "DESCRIPCIÓN", style=BOLD, align=CENTER, border=ALL_THIN)
    write_merged(row, *COL_UD,   "UD", style=BOLD, align=CENTER, border=ALL_THIN)
    write_merged(row, *COL_CANT, "CANTIDAD", style=BOLD, align=CENTER, border=ALL_THIN)
    write_merged(row, *COL_PU,   "P. UNITARIO", style=BOLD, align=CENTER, border=ALL_THIN)
    write_merged(row, *COL_TOT,  "TOTAL", style=BOLD, align=CENTER, border=ALL_THIN)
    row += 1

    total_presupuesto = 0.0

    # --- Datos ---
    for parent in parents:
        childs = _children_of_parent(items, parent)
        if not childs:
            continue

        # Sección (Padre)
        parent_name = nombres_padres.get(parent, f"SECCIÓN {parent}")
        
        write_merged(row, *COL_ITEM, parent, style=BOLD)
        write_merged(row, *COL_DESC, parent_name, style=BOLD)
        row += 1

        # Hijos
        for child in childs:
            if child not in df_datos_idx.index:
                continue
            item_row = df_datos_idx.loc[child]
            
            desc = str(item_row.get("Partida", "") or "")
            ud = str(item_row.get("cantidad tipo", "") or "")
            cant = float(pd.to_numeric(item_row.get("cantidad numero", ""), errors="coerce") or 0.0)
            punit_clp = float(precios_item.get(str(child), 0.0))
            punit = punit_clp / factor_conversion if factor_conversion > 0 else punit_clp
            total_child = punit * cant
            total_presupuesto += total_child

            write_merged(row, *COL_ITEM, child)
            write_merged(row, *COL_DESC, desc)
            write_merged(row, *COL_UD,   ud, align=CENTER)
            write_merged(row, *COL_CANT, cant, align=RIGHT, number_format=FMT_QTY)
            write_merged(row, *COL_PU,   punit, align=RIGHT, number_format=FMT_MONEY)
            write_merged(row, *COL_TOT,  total_child, align=RIGHT, number_format=FMT_MONEY)
            row += 1

    # --- HOJA COSTOS ---
    ws_resumen = wb.create_sheet("Costos")
    
    # Configurar anchos hoja Costos
    ws_resumen.column_dimensions["A"].width = 2
    for c_idx in range(2, 20):
        col_letter = get_column_letter(c_idx)
        ws_resumen.column_dimensions[col_letter].width = 5
        
    r_res = 2
    # Titulo Hoja 2
    write_merged(r_res, 2, 10, "RESUMEN DE COSTOS", style=BOLD_U, target_ws=ws_resumen)
    r_res += 2
    
    # Calculos
    monto_utilidad = total_presupuesto * (porc_utilidad / 100.0)
    total_neto = total_presupuesto + monto_utilidad
    monto_iva = total_neto * (porc_iva / 100.0)
    presupuesto_total = total_neto + monto_iva
    
    # 1. TOTAL COSTO DIRECTO
    write_merged(r_res, 2, 8, "TOTAL COSTO DIRECTO", style=BOLD, target_ws=ws_resumen)
    write_merged(r_res, 12, 5, total_presupuesto, style=BOLD, number_format=FMT_MONEY, align=RIGHT, target_ws=ws_resumen)
    r_res += 2 # Salto linea
    
    # 2. Utilidad
    label_util = f"UTILIDAD ({porc_utilidad}%)"
    write_merged(r_res, 2, 8, label_util, style=BOLD, target_ws=ws_resumen)
    write_merged(r_res, 12, 5, monto_utilidad, number_format=FMT_MONEY, align=RIGHT, target_ws=ws_resumen)
    r_res += 2 # Salto linea
    
    # 3. TOTAL NETO
    write_merged(r_res, 2, 8, "TOTAL NETO", style=BOLD, target_ws=ws_resumen)
    write_merged(r_res, 12, 5, total_neto, style=BOLD, number_format=FMT_MONEY, align=RIGHT, target_ws=ws_resumen)
    r_res += 2 # Salto linea
    
    # 4. IVA
    label_iva = f"IVA ({porc_iva}%)"
    write_merged(r_res, 2, 8, label_iva, style=BOLD, target_ws=ws_resumen)
    write_merged(r_res, 12, 5, monto_iva, number_format=FMT_MONEY, align=RIGHT, target_ws=ws_resumen)
    r_res += 2 # Salto linea
    
    # 5. PRESUPUESTO TOTAL
    write_merged(r_res, 2, 8, "PRESUPUESTO TOTAL", style=BOLD, target_ws=ws_resumen)
    write_merged(r_res, 12, 5, presupuesto_total, style=BOLD, number_format=FMT_MONEY, align=RIGHT, target_ws=ws_resumen)

    # --- Guardar ---
    if salida is None:
        salida = proyecto_dir / "presupuesto_detallado.xlsx"
    else:
        salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))
    return salida


# ---------------- Vista Streamlit ----------------
def render_crear_detallado():
    st.title("🧾 Crear Presupuesto Detallado")

    proyectos = _listar_proyectos(".")
    if not proyectos:
        st.info("No hay proyectos con `datos.csv` y `detalle.csv`.")
        return

    proyecto = st.selectbox("Proyecto", proyectos, index=0)

    # Inputs Generales
    colA, colB = st.columns(2)
    ubicacion = colA.text_input("Ubicación", value="")
    propietario = colB.text_input("Nombre Propietario", value="")
    
    # Inputs Porcentajes
    colC, colD = st.columns(2)
    porc_utilidad = colC.number_input("Utilidad (%)", min_value=0.0, value=10.0, step=0.1)
    porc_iva = colD.number_input("IVA (%)", min_value=0.0, value=19.0, step=0.1)

    # Detectar padres y pedir nombres
    try:
        # Forzar que Item sea string para no perder '1.00' a '1.0'
        df_datos = pd.read_csv(Path("presupuestos") / proyecto / "datos.csv", dtype={"Item": str})
        items = sorted(df_datos["Item"].astype(str).tolist(), key=_parse_key)
        parents = _collect_parents(items)
    except Exception as e:
        parents, items = [], []
        st.error(f"No se pudieron cargar ítems: {e}")

    if parents:
        st.caption("Nombra las secciones (ítems padre con hijos):")
        nombres_padres: Dict[str, str] = {}
        
        for p in parents:
            # No buscamos nada automáticamente, solo pedimos el nombre al usuario.
            default_val = f"SECCIÓN {p}"
            nombres_padres[p] = st.text_input(f"Nombre para {p}", value=default_val)
    else:
        nombres_padres = {}

    ruta_excel = Path("presupuestos") / proyecto / "presupuesto_detallado.xlsx"
    st.caption(f"Excel de salida: `{ruta_excel}`")

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("⚙️ Generar Excel", use_container_width=True):
            try:
                path = generar_excel_detallado(
                    proyecto=proyecto,
                    nombres_padres=nombres_padres,
                    ubicacion=ubicacion,
                    propietario=propietario,
                    porc_utilidad=porc_utilidad,
                    porc_iva=porc_iva
                )
                st.success(f"Excel generado: {path.name}")
            except Exception as e:
                st.error(f"Error al generar Excel: {e}")
                # Imprimir error completo a consola para debug
                import traceback
                traceback.print_exc()

    with c2:
        if st.button("👁️ Abrir Excel", use_container_width=True):
            if ruta_excel.exists():
                _abrir_en_sistema(ruta_excel)
                st.success("Abriendo Excel...")
            else:
                st.warning("Primero genera el Excel.")

    with c3:
        data_excel = None
        if ruta_excel.exists():
            with open(ruta_excel, "rb") as f:
                data_excel = f.read()

        st.download_button(
            "⬇️ Descargar Excel",
            data=data_excel if data_excel else b"",
            file_name=f"{proyecto}_presupuesto_detallado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=(data_excel is None),
            use_container_width=True,
        )

    st.caption(f"Estado: {'✅ Existe' if ruta_excel.exists() else '❌ No existe (se creará al generar)'}")
